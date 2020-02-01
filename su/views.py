from django.shortcuts import render
from django.http import HttpResponse
from django.contrib.auth import authenticate, login
from django.http import HttpResponseRedirect, HttpResponse
from django.contrib.auth.decorators import login_required
from django.contrib.auth import logout
from su.models import Coord,Student,Wear,Event,Wear_Student,Event_Student,DateMailStatus, Wear_Invalid_Students, Event_Invalid_Students
from su.forms import CoordUserForm,CoordUserProfileForm,UploadFileForm, WearForm,EventForm, WearFormEdit, EventFormEdit,ExcelUpload2,ExcelUpload3
from django.conf import settings
#addddddd
from django import forms
from django.shortcuts import render_to_response
from django.http import HttpResponseBadRequest, JsonResponse, HttpResponse
from django.template import RequestContext
import django_excel as excel
from datetime import datetime,date,timedelta
from django.core.mail import send_mail,send_mass_mail
from django.views.decorators.cache import cache_control
import xlsxwriter
from django.contrib.auth.models import User
import os
from django.core import mail
import traceback
import math
import openpyxl

try:
    import cStringIO as StringIO
except ImportError:
    import StringIO

####
#from django_cron import CronJobBase, Schedule
#To be triggered
from django.core.mail import EmailMultiAlternatives

def datechecker(gmid):
	try:
		item = Wear.objects.get(gm_id=gmid)
	except:
		try:
			item = Event.objects.get(gm_id=gmid)
		except:
			pass
			return HttpResponse("Error")
	c=date.today()
	d=0
	if (item.deadline2>c):     # coord reg open
		return 2
	elif (item.deadline >=c and item.deadline2<= c):   #student register/cancel  , coord reg closed
		return 1
	elif (item.deadline <c and item.date>=c):   #student reg/cancel closed
		return 3
	elif (item.date<c):    # coord spot signing upload
		return 4

def home(request):
	return render(request, 'su/home.html')

def about(request):
	return render(request, 'su/about.html')

def contact(request):
	return render(request, 'su/contact.html')

def su_item_sendmail1(request,gmid):
	itemid = request.GET.get('itemid')
	datemail = DateMailStatus.objects.get(date=datetime.now())
	try:
		item = Wear.objects.get(name=str(itemid))
		print(gmid)
		d = datetime.strptime(str(item.date), '%Y-%m-%d')
		e = date.strftime(d, "%d %B %Y")
		v = datetime.strptime(str(item.deadline), '%Y-%m-%d')
		h = date.strftime(v, "%d %B %Y")
		abcd=Wear_Student.objects.filter(gm_id=item.gm_id,status="Signed Up",mail="Not Sent")
		print e
		print h
		k=len(abcd)//99
		count=0
		for q in range(k+1):
			a=[]
			students=abcd[q*99:(q+1)*99]
			for j in students:
				a.append(str(j.user_id)+"@pilani.bits-pilani.ac.in")
				j.mail = "Sent"
				j.save()
			print a
			# smtp: send_mail(subject,message,from_email, fail_silently=True)
			subject, from_email = "[SU Portal] " + str(item.name), settings.EMAIL_HOST_USER
			text_content = 'Cancellation Mail'
			html_content = "<body><p>This is to inform you that you have been signed up for <strong> "+str(item.name)+"</strong>.</p> <p>In case you wish to cancel your signing, please visit <a href=http://su-bitspilani.org/ >SU Portal</a>, <strong> before 11:59 PM, " + h + "</strong>. Any requests made after the Cancellation Deadline will not be entertained. </p> </p><p>Thank you.</p><p>Students' Union 2018-19</p><p>Note: This mail is automatically generated via the Students' Union Portal. If you need technical assistance, contact SU Tech Team.</body>"
			msg = EmailMultiAlternatives(subject, text_content, from_email, cc = a, bcc=["f20171037@pilani.bits-pilani.ac.in"])
			msg.attach_alternative(html_content, "text/html")
			# print a
			try :
				connection = mail.get_connection(username=settings.EMAIL_HOST_USER, password=settings.EMAIL_HOST_PASSWORD, fail_silently=False)
				# connection.open()
				# print a
				messages = list()
				# msg.send(fail_silently=False)

				messages.append(msg)
				print a
				connection.send_messages(messages)
				count=count + len(a)
				print a
				datemail.mails = datemail.mails+len(a)
				datemail.save()
			except Exception as e:
				print(e)
				for j in students:
					j.mail = "Not Sent"
					j.save()
				left = len(abcd)-count
				data ={'is_taken': "Only "+str(count)+" mails were sent succesfully. " + str(left) +" mails are left to be send. Error" + str(e) }
				return JsonResponse(data)
		item.mails="Sent"
		item.save()
		print(gmid)
		data = {'is_taken': "All the mails ("+ str(len(abcd)) +") were sent succesfully"}
		return JsonResponse(data)
	except :
		pass
	try:
		item = Event.objects.get(name=str(itemid))
		print(gmid)
		d = datetime.strptime(str(item.date), '%Y-%m-%d')
		e = date.strftime(d, "%d %B %Y")
		v = datetime.strptime(str(item.deadline), '%Y-%m-%d')
		h = date.strftime(v, "%d %B %Y")
		abcd=Event_Student.objects.filter(gm_id=item.gm_id,status="Signed Up",mail="Not Sent")
		print e
		print h
		k=len(abcd)//99
		count=0
		for q in range(k+1):
			a=[]
			students=abcd[q*99:(q+1)*99]
			for j in students:
				a.append(str(j.user_id)+"@pilani.bits-pilani.ac.in")
				j.mail = "Sent"
				j.save()
			print a
			subject, from_email = "[SU Portal] " + str(item.name), 'su.pilani@gmail.com'
			text_content = 'Cancellation Mail'
			html_content = "<body><p>This is to inform you that you have been signed up for <strong> "+str(item.name)+"</strong>.</p> <p>In case you wish to cancel your signing, please visit <a href=http://su-bitspilani.org/>SU Portal</a>, <strong> before 11:59 PM, " + h +"</strong>. Any requests made after the Cancellation Deadline will not be entertained. </p> </p><p>Thank you.</p><p>Students' Union 2018-19</p><p>Note: This mail is automatically generated via the Students' Union Portal. If you need technical assistance, contact SU Tech Team.</body>"
			msg = EmailMultiAlternatives(subject, text_content, from_email, cc = a, bcc=["f20171037@pilani.bits-pilani.ac.in"])
			msg.attach_alternative(html_content, "text/html")
			print a
			try :
				msg.send(fail_silently=False)
				count=count + len(a)
				datemail.mails = datemail.mails+len(a)
				datemail.save()
			except :
				for j in students:
					j.mail = "Not Sent"
					j.save()
				left = len(abcd)-count
				data ={'is_taken': "Only "+str(count)+" mails were sent succesfully. " + str(left) +" mails are left to be send. Error" + str(e) }
				return JsonResponse(data)
		item.mails="Sent"
		item.save()
		print(gmid)
		data = {'is_taken': "All the mails ("+ str(len(abcd)) +") were sent succesfully"}
		return JsonResponse(data)
	except :
		print("here")
		data = {'is_taken': "Internal Error"}
		return JsonResponse(data) # sends the mail#send mails


def su_item_sendmail(request,gmid):#redirects to send mail page
	context_dict={}
	try:
		item = Wear.objects.get(gm_id=gmid)
		context_dict["item"]=item
		stud = Wear_Student.objects.filter(gm_id=gmid,status="Signed Up")
		registered=len(stud)
		print("here")
		context_dict["registered"]=registered
		if(DateMailStatus.objects.filter(date=datetime.now()).exists()):
			print("here")
			datemail = DateMailStatus.objects.get(date=datetime.now())
			context_dict["datemail"]=datemail
		else :
			print("here2")
			datemail = DateMailStatus.objects.create(date=datetime.now(),mails=0)
			context_dict["datemail"]=datemail
		mailsleft = 1000 - datemail.mails
		context_dict["mailsleft"]=mailsleft
		return render(request,'su/su_item_sendmail.html',context_dict)
	except:
		pass
	try:
		item = Event.objects.get(gm_id=gmid)
		context_dict["item"]=item
		stud = Event_Student.objects.filter(gm_id=gmid,status="Signed Up")
		registered=len(stud)
		print("here")
		context_dict["registered"]=registered
		if(DateMailStatus.objects.filter(date=datetime.now()).exists()):
			print("here")
			datemail = DateMailStatus.objects.get(date=datetime.now())
			context_dict["datemail"]=datemail
		else :
			print("here2")
			datemail = DateMailStatus.objects.create(date=datetime.now(),mails=0)
			context_dict["datemail"]=datemail
		mailsleft = 1000 - datemail.mails
		context_dict["mailsleft"]=mailsleft
		return render(request,'su/su_item_sendmail.html',context_dict)
	except:
		pass
		return HttpResponseRedirect('/su/')
	return render(request,'su/su_item_sendmail.html',context_dict) # redirects to mail sending


def index(request): #Done
	# print('Hello1')
	if not request.user.is_authenticated():
		return render(request, 'su/index.html')
	else:
		try:
			coord=Coord.objects.get(user=request.user)
			wear= Wear.objects.filter(cg_id=coord.cg_id).order_by('-date')
			event= Event.objects.filter(cg_id=coord.cg_id).order_by('-date')
			context_dict={'coord':coord,'wear':wear,'event':event}
			return render(request, 'su/index.html',context_dict)
		except Coord.DoesNotExist:
			context_dict={}
			pass
		try:
			# print(request.user.email[:-25])
			stud= Student.objects.get(user_id=str(request.user.email[:-25]))
			# print(stud)
			context_dict={'student':stud}
			# print(str(request.user))
			# print('Hello')
			return render(request, 'su/index.html',context_dict)
		except Student.DoesNotExist:
			context_dict={}
			# print('Hello3')
			pass
		return render(request, 'su/index.html',context_dict)

def su_login(request): #Done
	context_dict={}
    	if request.method == 'POST':
		username = request.POST.get('username')
		password = request.POST.get('password')

		user = authenticate(username=username, password=password)

		if user and user.is_superuser:
		    	if user.is_active:
				login(request, user)
				return HttpResponseRedirect('/su/')
		    	else:
				return HttpResponse("Your su account is disabled.")
		else:

			context_dict['invalid']="Invalid login details supplied."
		    	return render(request, 'su/su_login.html',context_dict)

    	else:

		return render(request, 'su/su_login.html')




@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def su_register(request): #Done, admin registers coord
	if request.user.is_superuser:
		registered = False
		if request.method == 'POST':
			user_form = CoordUserForm(data=request.POST)
			profile_form = CoordUserProfileForm(data=request.POST)
			if user_form.is_valid() and profile_form.is_valid():
				user = user_form.save(commit=False)
				user.set_password(user.password)
				user.is_staff=True
				profile = profile_form.save(commit=False)
				profile.status="Active"
				profile.date=datetime.now()
				profile.reg_by=str(request.user)
				user.save()
				profile.user = user
				profile.save()
				registered = True
			else:
				print user_form.errors
				print profile_form.errors

		else:
			user_form = CoordUserForm()
			profile_form = CoordUserProfileForm()
		return render(request,'su/su_register.html',{'user_form': user_form, 'profile_form': profile_form, 'registered': registered} )
	else :
		return HttpResponseRedirect('/su/su/login/')


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def su_iteminfo(request,gmid): # DOne
	if request.user.is_superuser:
		#e=datechecker(gmid)
		e=2
		context_dict = {}
		try:
			item = Wear.objects.get(gm_id=gmid)
			context_dict['wear'] = item
			context_dict['e'] = e
			return render(request, 'su/su_iteminfo.html', context_dict)
		except Wear.DoesNotExist:
			pass
		try:
			item = Event.objects.get(gm_id=gmid)
			context_dict['event'] = item
			context_dict['e'] = e
			return render(request, 'su/su_iteminfo.html', context_dict)
		except Event.DoesNotExist:
			pass
			return HttpResponseRedirect('/su/')
		return render(request, 'su/su_iteminfo.html', context_dict)
	else :
		return HttpResponseRedirect('/su/su/login/')


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def su_item_list(request): #Done
	if request.user.is_superuser:
		wear_list=Wear.objects.order_by('-date')[:] # reverse chronological
		event_list=Event.objects.order_by('-date')[:] # reverse chronological
		coord_list=Coord.objects.all()
		context_dict={"wear":wear_list,"event":event_list,"coord":coord_list}
		return render(request, 'su/su_itemlist.html',context_dict)
	else :
		return HttpResponseRedirect('/su/su/login/')


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def su_student_table(request,gmid): #Done
	if request.user.is_superuser :
		try:
			item=Wear.objects.get(gm_id=gmid)
			stud = Wear_Student.objects.filter(gm_id=gmid)
			registered=len(stud.filter(status="Signed Up"))
			out=len(stud.filter(status="Opted Out"))
			context_dict={"stud":stud,"reg":registered,"out":out,"gmid":gmid,"item":item}
			return render(request, 'su/dynamic_table.html',context_dict)
		except:
			pass
		try:
			item=Event.objects.get(gm_id=gmid)
			stud = Event_Student.objects.filter(gm_id=gmid)
			registered=len(stud.filter(status="Signed Up"))
			out=len(stud.filter(status="Opted Out"))
			context_dict={"stud":stud,"reg":registered,"out":out,"gmid":gmid,"item":item}
			return render(request, 'su/dynamic_table.html',context_dict )
		except:
			pass
		return HttpResponseRedirect("/su")

	elif request.user.is_staff :
		try:
			item=Wear.objects.get(gm_id=gmid)
			e=2
			coord = Coord.objects.get(cg_id=item.cg_id.cg_id)
			if request.user == coord.user:
				item=Wear.objects.get(gm_id=gmid)
				stud = Wear_Student.objects.filter(gm_id=gmid)
				registered=len(stud.filter(status="Signed Up"))
				out=len(stud.filter(status="Opted Out"))
				context_dict={"stud":stud,"reg":registered,"out":out,"gmid":gmid,"item":item}
				return render(request, 'su/dynamic_table.html',context_dict)
		except:
			pass
		try:
			item=Event.objects.get(gm_id=gmid)
			e=2
			coord = Coord.objects.get(cg_id=item.cg_id.cg_id)
			if request.user == coord.user:
				item=Event.objects.get(gm_id=gmid)
				stud = Event_Student.objects.filter(gm_id=gmid)
				registered=len(stud.filter(status="Signed Up"))
				out=len(stud.filter(status="Opted Out"))
				context_dict={"stud":stud,"reg":registered,"out":out,"gmid":gmid,"item":item}
				return render(request, 'su/dynamic_table.html',context_dict)
		except:
			pass
			return HttpResponseRedirect("/su")
	else:
		return HttpResponseRedirect('/su/su/login/')


def su_coord_active(request,cgid):   ##Done
	if request.user.is_superuser:
		a=Coord.objects.get(cg_id=cgid)
		b=User.objects.get(username=a.user)
		try:
			c=Wear.objects.filter(cg_id=cgid)
			for i in c:
				i.status="Active"
				i.save()
		except Wear.DoesNotExist:
			pass
		try:
			c=Event.objects.filter(cg_id=cgid)
			for i in c:
				i.status="Active"
				i.save()
		except Event.DoesNotExist:
			pass
		b.is_staff=True
		b.is_active=True
		a.status="Active"
		a.save()
		b.save()
		return HttpResponseRedirect("/su/su/itemlist/")
	else :
		return HttpResponseRedirect('/su/su/login/')



def su_coord_inactive(request,cgid):	##Done
	if request.user.is_superuser:
		a=Coord.objects.get(cg_id=cgid)
		b=User.objects.get(username=a.user)
		try:
			c=Wear.objects.filter(cg_id=cgid)
			for i in c:
				i.status="Inactive"
				i.save()
		except Wear.DoesNotExist:
			pass
		try:
			c=Event.objects.filter(cg_id=cgid)
			for i in c:
				i.status="Inactive"
				i.save()
		except Event.DoesNotExist:
			pass
		b.is_staff=False
		b.is_active=False
		a.status="Inactive"
		a.save()
		b.save()
		return HttpResponseRedirect("/su/su/itemlist/")
	else :
		return HttpResponseRedirect('/su/su/login/')



def export_data(request, gmid): #Done
	if request.user.is_staff or not request.user.is_superuser:
		if request.method == "POST":
			try :
				item =Wear.objects.get(gm_id=gmid)
				bh = request.POST.get('bhawan')
				if bh=="All":
					c = Wear_Student.objects.filter(gm_id=gmid,status="Signed Up").order_by('student_id')
				else :
					c = Wear_Student.objects.filter(gm_id=gmid,status="Signed Up",bhawan=bh).order_by('student_id')
				a= Wear.objects.get(gm_id=gmid)
				b=[]
				for stu in c:
					b.append([stu.student_id,stu.name, stu.bhawan,stu.room,stu.meal,str(stu.price)])
				output = StringIO.StringIO()
				workbook = xlsxwriter.Workbook(output)
				worksheet = workbook.add_worksheet()
				worksheet.set_column('A:A', 20) # ID
				worksheet.set_column('B:B', 25) # name
				worksheet.set_column('C:C', 15) # Bhawan
				worksheet.set_column('D:D', 15) # Room
				worksheet.set_column('E:E', 15) # size
				worksheet.set_column('F:F', 15) # price
				bold = workbook.add_format({'bold': 1})
				worksheet.write('A1', 'BITS ID', bold)
				worksheet.write('B1', 'Name', bold)
				worksheet.write('C1', 'Bhawan', bold)
				worksheet.write('D1', 'Room No.', bold)
				worksheet.write('E1', 'Size', bold)
				worksheet.write('F1', 'Amount', bold)
				row = 1
				col = 0
				for i in b:
					worksheet.write_string(row, col,i[0] )
					worksheet.write_string(row, col + 1, i[1])
					worksheet.write_string(row, col + 2, i[2])
					worksheet.write_string(row, col + 3, i[3])
					worksheet.write_string(row, col + 4, i[4])
					worksheet.write_string(row, col + 5, i[5])
					row += 1
				workbook.close()
				# Downlaod list
				output.seek(0)
				response = HttpResponse(output.read(), content_type="application/ms-excel")
				response['Content-Disposition'] = 'attachment; filename="' + a.name+'_'+bh+'_itemlist.xlsx"'
				return response
			except Wear.DoesNotExist:
				pass
			try :
				item =Event.objects.get(gm_id=gmid)
				bh = request.POST.get('bhawan')
				if bh=="All":
					c = Event_Student.objects.filter(gm_id=gmid,status="Signed Up").order_by('student_id')
				else :
					c = Event_Student.objects.filter(gm_id=gmid,status="Signed Up",bhawan=bh).order_by('student_id')
				a= Event.objects.get(gm_id=gmid)

				b=[]
				for stu in c:
					b.append([stu.student_id, stu.name, stu.bhawan,stu.room,str(stu.price)])
				print(b)
				output = StringIO.StringIO()
				workbook = xlsxwriter.Workbook(output)
				worksheet = workbook.add_worksheet()
				worksheet.set_column('A:A', 20) # ID
				worksheet.set_column('B:B', 25) # Name
				worksheet.set_column('C:C', 15) # Bhawan
				worksheet.set_column('D:D', 15) # Room No.
				worksheet.set_column('E:E', 15) # Price
				bold = workbook.add_format({'bold': 1})
				worksheet.write('A1', 'BITS ID', bold)
				worksheet.write('B1', 'Name', bold)
				worksheet.write('C1', 'Bhawan', bold)
				worksheet.write('D1', 'Room No.', bold)
				worksheet.write('E1','Amount', bold)
				row = 1
				col = 0
				for i in b:
					worksheet.write_string  (row, col,i[0] )
					worksheet.write_string(row, col + 1, i[1] )
					worksheet.write_string  (row, col + 2,i[2] )
					worksheet.write_string  (row, col+3,i[3] )
					worksheet.write_string(row, col + 4, i[4] )
					row += 1
				workbook.close()
				# Downlaod list
				output.seek(0)
				response = HttpResponse(output.read(), content_type="application/ms-excel")
				response['Content-Disposition'] = 'attachment; filename="' + a.name+'_'+bh+'_itemlist.xlsx"'
				return response
			except Event.DoesNotExist:
				pass
		else:
			return HttpResponseRedirect("/su")

	else :
		return HttpResponseRedirect('/su/')



def coord_login(request):#DOne
	context_dict={}
	if request.method == 'POST':
		username = request.POST.get('username')
		password = request.POST.get('password')

        	user = authenticate(username=username, password=password)

		if user and user.is_staff:
		    	if user.is_active:                                                 #add status  choice here
		        	login(request, user)
		        	return HttpResponseRedirect('/su/')
		    	else:
		        	return HttpResponse("Your su account is disabled.")
		else:
			context_dict['invalid']="Invalid login details supplied."
		    	print "Invalid login details: {0}, {1}".format(username, password)
		    	return render(request,'su/coord_login.html', context_dict)

    	else:

       		return render(request, 'su/coord_login.html', {})


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def coord_item_register(request): #DOne
	if request.user.is_staff and not request.user.is_superuser:
		done=0
		if request.method == 'POST':
			try:
				form1= WearForm(request.POST,request.FILES)
				if form1.is_valid() :
					a=form1.save(commit=False)
					a.date=request.POST.get('itemdate1')
					# a.reg_date=datetime.now()
					d=Coord.objects.get(user=request.user)
					a.cg_id=d
					a.save()
					qw=Wear.objects.get(gm_id = a.gm_id)
					b=qw.date
					c=timedelta(days=2)
					e=timedelta(days=4)
					a.deadline=b-c
					a.deadline2=b-e
					a.status="Active"
					a.save()
					if 'image' in request.FILES :
						a.image = request.FILES['image']
						a.save()
						done=1
			except:
				pass
			try:
				form2= EventForm(request.POST,request.FILES)
				if form2.is_valid() :
					a=form2.save(commit=False)
					a.date=request.POST.get('itemdate2')
					a.reg_date=datetime.now()
					d=Coord.objects.get(user=request.user)
					a.cg_id=d
					a.save()
					qw=Event.objects.get(gm_id = a.gm_id)
					b=qw.date
					c=timedelta(days=2)
					e=timedelta(days=4)
					a.deadline=b-c
					a.deadline2=b-e
					a.status="Active"
					a.save()
					if 'image' in request.FILES :
						a.image = request.FILES['image']
						a.save()
						done=1
			except:
				pass
		# else:
		# 		print form1.errors
		# 		print form2.errors
		else:
			form1= WearForm()
			form2= EventForm()
		return render(request, 'su/coord_item_register.html', {'form1':form1,'form2':form2,'done':done})
	else :
		return HttpResponseRedirect('/su/coord/login/')


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def coord_upload(request,gmid): # Finally Done!
	if request.user.is_staff and not request.user.is_superuser:
		try :
			item =Wear.objects.get(gm_id=gmid)
			form = ExcelUpload2(instance=item)
			#e=datechecker(gmid)
			e=2
			coord = Coord.objects.get(cg_name=item.cg_id)
			if (request.user == coord.user):
				if request.method == 'POST' and request.FILES:
					b=request.FILES['excel']
					filename=str(b.name)
					if filename.endswith('.xls') or filename.endswith('.xlsx') or filename.endswith('.csv'):
						a = Wear.objects.get(gm_id=gmid)
						d=Wear.objects.filter(gm_id=gmid)[0]
						form = ExcelUpload2(request.POST,request.FILES,instance=item)
						if form.is_valid():
					   		photo=form.save(commit=False)
							if 'excel' in request.FILES :
								# print('Hullohulka')
								def choice_func(row):
									row[0]=row[0].upper()
									a=row[0]
									a=str(a).upper()[:] #Change to 13 if "P" is too be included
									row[1]=row[1].upper()
									r=row[1] # size
									row[2]=int(row[2])
									p= row[2] # price field
									try :
										b=Student.objects.get(bits_id=str(a))
										smailid=b.user_id
										sname=b.name
										sbhawan=b.bhawan
										sroom=b.room_no
										try :
											gs=Wear_Student.objects.get(student_id=str(a),gm_id=gmid)
											return None
										except :
											pass
										row.append(smailid)
										row.append(sname)
										row.append(sbhawan)
										row.append(sroom)

									except :
									    invalid_item, created = Wear_Invalid_Students.objects.get_or_create(
									        student_id=str(row[0]),gm_id=item,meal=str(row[1]).lower()
									        )
									    invalid_item.save()
									    return None
										# if a[4]=="H" or a[4]=="P":
										# 	smailid=a[4]+a[0:4]+a[8:12]
										# 	sname="User"
										# 	sbhawan="Not Specified"
										# 	sroom="Not Specified"
										# 	row.append(smailid)
										# 	row.append(sname)
										# 	row.append(sbhawan)
										# 	row.append(sroom)
										# else :
										# 	smailid="f"+a[0:4]+a[8:12]
										# 	print(smailid)
										# 	sname="User"
										# 	sbhawan="Not Specified"
										# 	sroom="Not Specified"
										# 	row.append(smailid)
										# 	row.append(sname)
										# 	row.append(sbhawan)
										# 	row.append(sroom)
										# row[1]=str(row[1]).upper()
									row.append("Signed Up")
									row.append(d)
									# row.append(r)
									# row.append(p)
									return row

								files=request.FILES['excel']
								files.save_to_database(
								model=Wear_Student,
								initializer=choice_func,
								mapdict=[ 'student_id', 'meal', 'price','user_id','name','bhawan', 'room','status','gm_id']
						    	) # maps a row of excel sheet to save the data into database
					    		# photo.excel = files
								photo.save()
								print("GOT IT!")
								return HttpResponseRedirect("/su/stats/"+gmid)
							else :
								invalid="No file uploaded."
								return render(request, 'su/coord_upload.html', {'form': form,'item':item,"e":e,"invalid":invalid})
						else:
							print form.errors
							invalid="Invalid File type."
							return render(request, 'su/coord_upload.html', {'form': form,'item':item,"e":e,"invalid":invalid})
					else :
						invalid="Unsupported File type."
						return render(request, 'su/coord_upload.html', {'form': form,'item':item,"e":e,"invalid":invalid})

				else:

					form = ExcelUpload2(instance=item)

				return render(request, 'su/coord_upload.html', {'form': form,'item':item,"e":e})

			else:
				return HttpResponseRedirect("/su")
		except Wear.DoesNotExist:
			pass
		try :
			item =Event.objects.get(gm_id=gmid)
			form = ExcelUpload3(instance=item)
			#e=datechecker(gmid)
			e=2
			coord = Coord.objects.get(cg_name=item.cg_id)
			if (request.user == coord.user):
				# if (e==2 or e==4):
				if request.method == 'POST' and request.FILES:
					b=request.FILES['excel']
					filename=str(b.name)
					if filename.endswith('.xls') or filename.endswith('.xlsx') or filename.endswith('.csv'):
						a = Event.objects.get(gm_id=gmid)
						d=Event.objects.filter(gm_id=gmid)[0]
						m = a.meal

						form = ExcelUpload3(request.POST,request.FILES,instance=item)
						if form.is_valid():
					   		photo=form.save(commit=False)
							if 'excel' in request.FILES :
								def choice_func(row):
									a=row[0]
									a=str(a).upper()[:]
									row[1]=int(row[1])
									p= row[1] # price field, no size needed
									try :
										b=Student.objects.get(bits_id=str(a))
										smailid=b.user_id
										sname=b.name
										sbhawan=b.bhawan
										sroom=b.room_no
										try :
											gs=Event_Student.objects.get(student_id=str(a),gm_id=gmid)
											return None
										except:
											pass
										row.append(smailid)
										row.append(sname)
										row.append(sbhawan)
										row.append(sroom)
									except :
										invalid_item, created = Event_Invalid_Students.objects.get_or_create(
									        student_id=str(row[0]),gm_id=item,meal="Workshop/FoodStall/ProfShow"
									        )
										invalid_item.save()
										return None
										# if a[4]=="H" or a[4]=="P":
										# 	smailid=a[4]+a[0:4]+a[8:12]
										# 	sname="User"
										# 	sbhawan="Not Specified"
										# 	sroom="Not Specified"
										# 	row.append(smailid)
										# 	row.append(sname)
										# 	row.append(sbhawan)
										# 	row.append(sroom)
										# else :
										# 	smailid="f"+a[0:4]+a[8:12]
										# 	sname="User"
										# 	sbhawan="Not Specified"
										# 	sroom="Not Specified"
										# 	row.append(smailid)
										# 	row.append(sname)
										# 	row.append(sbhawan)
										# 	row.append(sroom)
									row.append("Signed Up")
									row.append(d)
									return row
								files=request.FILES['excel']
								files.save_to_database(
								model=Event_Student,
								initializer=choice_func,
								mapdict=[ 'student_id','price','user_id','name','bhawan', 'room','status','gm_id']
							    	)
						    	# photo.excel = files
								photo.save()
								print("Got Event!")
								return HttpResponseRedirect("/su/stats/"+gmid)
						else:
							invalid = "No File Uploaded."
							return render(request, 'su/coord_upload.html', {'form': form,'item':item,"e":e,"invalid":invalid})
					else :
						print form.errors
						invalid="Unsupported File type."
						return render(request, 'su/coord_upload.html', {'form': form,'item':item,"e":e,"invalid":invalid})

				else:

					form = ExcelUpload3(instance=item)
				return render(request, 'su/coord_upload.html', {'form': form,'item':item,"e":e})

			else:
				return HttpResponseRedirect("/su")
		except Event.DoesNotExist:
			pass
			return HttpResponseRedirect("/su")
		except:
			#invalid="The following error occured : \n" + str(e1)
			return HttpResponse(traceback.format_exc())
			return render(request, 'su/coord_upload.html', {'form': form,'item':item,"e":e,"invalid":invalid})
	else :
		return HttpResponseRedirect('/su/coord/login/')


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def coord_student_register(request,gmid): # Incomplete, Not to be Done for now, add code for Event as well when implementing
	if request.user.is_staff:
		try :
			item =Wear.objects.get(gm_id=gmid)
			form = ExcelUpload(instance=item)
			#e=datechecker(gmid)
			e=2
			coord = Coord.objects.get(cg_name=item.cg_id)
			if request.user == coord.user or request.user.is_superuser:
				done=0
				if e==2:
					if request.method == 'POST':
						form = CoordStudentRegForm(request.POST)
						if form.is_valid():
							photo=form.save(commit=False)
							photo.student_id=photo.student_id.upper()
							try :
								gs=Wear_Student.objects.get(student_id=photo.student_id,gm_id=gmid)
								invalid="Student already registered."
								return render(request, 'su/coord_student_register.html',{'form': form,'done':done,'item':item,'e':e,"invalid":invalid})
							except Wear_Student.DoesNotExist:
								pass
							try :
								gs=Event_Student.objects.get(student_id=photo.student_id,gm_id=gmid)
								invalid="Student already registered."
								return render(request, 'su/coord_student_register.html',{'form': form,'done':done,'item':item,'e':e,"invalid":invalid})
							except Wear_Student.DoesNotExist:
								pass
							photo.status="Signed Up"
							photo.meal=request.POST.get('mealtype')
							a = photo.student_id
							if (a[4].upper()=="H" or a[4].upper()=="P"):
								if (int(a[0:4])<2017):
									photo.user_id=a[4]+a[0:4]+a[9:12]
								else :
									photo.user_id=a[4]+a[0:4]+a[8:12]
							else :
								if (int(a[0:4])<2017):
									photo.user_id='f'+a[0:4]+a[9:12]
								else :
									photo.user_id='f'+a[0:4]+a[8:12]
							photo.user_id = photo.user_id.lower()
							try :
								d=Student.objects.get(bits_id=a)
								photo.room=d.room_no
								photo.bhawan=d.bhawan
								photo.name=d.name
								b=Wear.objects.filter(gm_id=gmid)[0]
								photo.gm_id=b
								photo.save()
								done=1
							except Student.DoesNotExist:
								# pass
								invalid="Invalid ID"
								return render(request, 'su/coord_student_register.html',{'form': form,'done':done,'item':item,'e':e,"invalid":invalid})

						else:
							print form.errors
					else:
						form = CoordStudentRegForm()
						item = Wear.objects.get(gm_id=gmid)
					return render(request, 'su/coord_student_register.html', {'form': form,'done':done,'item':item,'e':e})
				else:
					return render(request, 'su/coord_student_register.html', {'item':item,'e':e})
			else:
				return HttpResponseRedirect("/su")
		except Wear.DoesNotExist:
			pass
			return HttpResponseRedirect("/su")

	else :
		return HttpResponseRedirect('/su/coord/login/')


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def coord_view_item(request,gmid): # Done
	if request.user.is_staff and not request.user.is_superuser:
		try :
			item =Wear.objects.get(gm_id=gmid)
			#e=datechecker(gmid)
			e=2
			coord = Coord.objects.get(cg_name=item.cg_id)
			if request.user == coord.user:
				context_dict = {}
			    	try:
					context_dict['wear'] = item
					context_dict['e']=e
			    	except Wear.DoesNotExist:
					pass
			    	return render(request, 'su/coord_iteminfo.html', context_dict)
			else:
				return HttpResponseRedirect("/su")
		except Wear.DoesNotExist:
			pass
		try:
			item =Event.objects.get(gm_id=gmid)
			#e=datechecker(gmid)
			e=2
			coord = Coord.objects.get(cg_name=item.cg_id)
			if request.user == coord.user:
				context_dict = {}
			    	try:
					context_dict['event'] = item
					context_dict['e']=e
			    	except Wear.DoesNotExist:
					pass
			    	return render(request, 'su/coord_iteminfo.html', context_dict)
			else:
				return HttpResponseRedirect("/su")
		except Event.DoesNotExist:
			pass
			return HttpResponseRedirect("/su")
	else :
		return HttpResponseRedirect('/su/coord/login/')


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def coord_item_edit(request,gmid):  # Done, not implemented for now
	if request.user.is_staff and not request.user.is_superuser:
		try :
			item =Wear.objects.get(gm_id=gmid)
			#e=datechecker(gmid)
			e=2
			coord = Coord.objects.get(cg_name=item.cg_id)
			if request.user == coord.user:
				done=0
				inst = Wear.objects.get(gm_id=gmid)
				if request.method == 'POST':
					form = WearFormEdit(request.POST, instance=inst)
					if form.is_valid():
						item=form.save(commit=False)
						item.save()
						done=1
				   	else:
						print form.errors
				else:
					form = WearFormEdit(instance=inst)
				return render(request, 'su/coord_item_edit.html', {'form': form,'done':done,'item':item})
			else:
				return HttpResponseRedirect("/su")
		except Wear.DoesNotExist:
			pass
		try :
			item =Event.objects.get(gm_id=gmid)
			#e=datechecker(gmid)
			e=2
			coord = Coord.objects.get(cg_name=item.cg_id)
			if request.user == coord.user:
				done=0
				inst = Event.objects.get(gm_id=gmid)
				if request.method == 'POST':
					form = EventFormEdit(request.POST, instance=inst)
					if form.is_valid():
						item=form.save(commit=False)
						item.save()
						done=1
				   	else:
						print form.errors
				else:
					form = EventFormEdit(instance=inst)
				return render(request, 'su/coord_item_edit.html', {'form': form,'done':done,'item':item})
			else:
				return HttpResponseRedirect("/su")
		except Wear.DoesNotExist:
			pass
			return HttpResponseRedirect("/su")
	else :
		return HttpResponseRedirect('/su/coord/login/')

# Coordinator do not have the permissions to cancel/activate items

def su_item_inactive(request,gmid):  # done
	if request.user.is_superuser:
		try:
			a=Wear.objects.get(gm_id=gmid)
			a.status="Inactive"
			a.save()
			return HttpResponseRedirect("/su/su/item/"+gmid)
		except Wear.DoesNotExist:
			pass
		try:
			a=Event.objects.get(gm_id=gmid)
			a.status="Inactive"
			a.save()
			return HttpResponseRedirect("/su/su/item/"+gmid)
		except Event.DoesNotExist:
			pass
		return HttpResponseRedirect("/su/su/item/"+gmid)
	else :
		return HttpResponseRedirect('/su/su/login/')


def su_item_active(request,gmid):	# Done
	if request.user.is_superuser:
		try:
			a=Wear.objects.get(gm_id=gmid)
			a.status="Active"
			a.save()
			return HttpResponseRedirect("/su/su/item/"+gmid)
		except Wear.DoesNotExist:
			pass
		try:
			a=Event.objects.get(gm_id=gmid)
			a.status="Active"
			a.save()
			return HttpResponseRedirect("/su/su/item/"+gmid)
		except Event.DoesNotExist:
			pass
		return HttpResponseRedirect("/su/su/item/"+gmid)
	else :
		return HttpResponseRedirect('/su/su/login/')

@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def student_upcoming_items(request): # Done
	if request.user.is_authenticated() and not request.user.is_staff:
		c = date.today()
		wear_list=Wear.objects.filter(status="Active").order_by('-date')[:]
		event_list=Event.objects.filter(status="Active").order_by('-date')[:]
		context_dict={"wear":wear_list,"event":event_list}
		return render(request, 'su/student_itemlist.html',context_dict)
	else :
		return HttpResponseRedirect("/su/")


@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def student_item_register(request, gmid):		# Done
	context_dict={}
	context_dict['gmid'] = gmid
	if request.user.is_authenticated() and not request.user.is_staff:
		try:
			item = Wear.objects.get(gm_id=gmid,status="Active")
			coord = Coord.objects.get(cg_id=item.cg_id.cg_id)
			try :
				a = Wear_Student.objects.get(gm_id=gmid,user_id=str(request.user))
				b = a.meal
				c = a.price
				context_dict['student_price'] = c
				context_dict['meal'] = b
				context_dict['student']= a
			except Wear_Student.DoesNotExist:
				pass
			e=datechecker(gmid)
			print(e)
			context_dict['coord'] = coord
			context_dict['wear'] = item
			context_dict['e']=e
			return render(request, 'su/student_iteminfo.html', context_dict)
		except Wear.DoesNotExist:
			pass
		try:
			item = Event.objects.get(gm_id=gmid,status="Active")
			coord = Coord.objects.get(cg_id=item.cg_id.cg_id)
			try :
				a = Event_Student.objects.get(gm_id=gmid,user_id=str(request.user))
				b = a.meal
				c = a.price
				context_dict['student_price'] = c
				context_dict['meal'] = b
				context_dict['student']= a
			except Event_Student.DoesNotExist:
				pass
			e=datechecker(gmid)
			context_dict['coord'] = coord
			context_dict['event'] = item
			context_dict['e']=e
			return render(request, 'su/student_iteminfo.html', context_dict)
		except Wear.DoesNotExist:
			pass

	else:
		return render(request, 'su/student_iteminfo.html', context_dict)
		return HttpResponseRedirect("/soc/login/google-oauth2/?next=/su/student/item/"+gmid) # oauth2


def student_item_register4(request, gmid): #Done
	if request.user.is_authenticated() and not request.user.is_staff:
		try:
			item = Wear.objects.get(gm_id=gmid,status="Active")
			d=datechecker(gmid)
			if (d==1 or d==2) :
				print("here2")
				a = Wear.objects.filter(gm_id=gmid)[0]
				d= Student.objects.get(user_id=str(request.user))
				try :
					print("here100")
					b=Wear_Student.objects.get(gm_id=gmid,user_id=str(request.user))
					b.meal=request.POST.get('size')
					b.status="Signed Up"
					b.price=a.price
					b.save()
				except Wear_Student.DoesNotExist:
					print(a.price)
					Wear_Student.objects.create(gm_id=a,user_id=str(request.user),student_id=str(d.bits_id),status="Signed Up", room=d.room_no ,bhawan=d.bhawan,name=d.name, meal=request.POST.get('size'),price=a.price)

				return HttpResponseRedirect("/su/student/item/"+gmid+"/")
			else :
				return HttpResponseRedirect("/su/student/item/"+gmid+"/")
		except Wear.DoesNotExist:
			print("here4")
			pass
		try:
			item = Event.objects.get(gm_id=gmid,status="Active")
			d=datechecker(gmid)
			if (d==1 or d==2) :
				a = Event.objects.filter(gm_id=gmid)[0]
				d= Student.objects.get(user_id=str(request.user))
				try :
					b=Event_Student.objects.get(gm_id=gmid,user_id=str(request.user))
					b.meal=a.meal
					b.status="Signed Up"
					b.price=a.price
					b.save()
				except Event_Student.DoesNotExist:
					Event_Student.objects.create(gm_id=a,user_id=str(request.user),student_id=str(d.bits_id),status="Signed Up", room=d.room_no ,bhawan=d.bhawan,name=d.name,price=a.price,meal=a.meal)

				return HttpResponseRedirect("/su/student/item/"+gmid+"/")
			else :
				return HttpResponseRedirect("/su/student/item/"+gmid+"/")
		except Event.DoesNotExist:
			pass
			return HttpResponseRedirect("/su/")
	else :
		return HttpResponseRedirect("/su/")

def student_item_cancel(request, gmid):	#Done
	if request.user.is_authenticated() and not request.user.is_staff:
		try:
			item = Wear.objects.get(gm_id=gmid,status="Active")
			d=datechecker(gmid)
			if d==1 :
				try :
					a=Wear_Student.objects.get(gm_id=gmid,user_id=str(request.user))
					a.status="Opted Out"
					a.save()
				except Wear_Student.DoesNotExist:
					pass
				return HttpResponseRedirect("/su/student/item/"+gmid+"/")
			else :
				return HttpResponseRedirect("/su/student/item/"+gmid+"/")
		except Wear.DoesNotExist:
			pass
		try:
			item = Event.objects.get(gm_id=gmid,status="Active")
			d=datechecker(gmid)
			if d==1 :
				try :
					a=Event_Student.objects.get(gm_id=gmid,user_id=str(request.user))
					a.status="Opted Out"
					a.save()
				except Event_Student.DoesNotExist:
					pass
				return HttpResponseRedirect("/su/student/item/"+gmid+"/")
			else :
				return HttpResponseRedirect("/su/student/item/"+gmid+"/")
		except Event.DoesNotExist:
			pass
			return HttpResponseRedirect("/su/")
	else :
		return HttpResponseRedirect("/su/")

@login_required
def user_logout(request): #Done
	logout(request)
	return HttpResponseRedirect('/su/')

def import_data(request): #Done
	if request.user.is_superuser:
		done=0
	    	if request.method == "POST":
			form = UploadFileForm(request.POST,request.FILES)
			if form.is_valid():
				def choice_func(row):
					a=row[0].upper()
					if a[4]=="H" or a[4]=="P":
						if (int(a[0:4])<2017):
							b=a[4]+a[0:4]+a[9:12] # "P" and "0" Not included
						else :
							b=a[4]+a[0:4]+a[8:12] # "P" Not included , 0 included
					else :
						if (int(a[0:4])<2017):
							b='f'+a[0:4]+a[9:12] # "P" and "0" Not included
						else :
							b='f'+a[0:4]+a[8:12] # "P" Not included , 0 included
					row.append(b.lower())

				    	return row
				request.FILES['file'].save_to_database(
				model=Student,
				initializer=choice_func,
				mapdict=['bits_id','name','bhawan','room_no','user_id',]
			   	)
				done=1

		else:
			form = UploadFileForm()
	    	return render(request,'su/upload_form.html',{'form': form,'d':done})
	else :
		return HttpResponseRedirect("/su/su/login/")


def export(request):
	student = Student.objects.all()
	b=[]
	for i in student:
		b.append(i.name)
	workbook = xlsxwriter.Workbook('media/'+'uploaded_student_list.xlsx')
	worksheet = workbook.add_worksheet()
	worksheet.set_column('A:A', 20)
	bold = workbook.add_format({'bold': 1})
	worksheet.write('A1', 'User ID', bold)
	row = 1
	col = 0
	for i in b:
		worksheet.write_string  (row, col,i )
		row += 1
	workbook.close()
	return (b)
	return HttpResponseRedirect('media/'+'uploaded_student_list.xlsx')


def coord_invalid_ids(request,gmid):
    if request.user.is_superuser:
        try:
			item=Wear.objects.get(gm_id=gmid)
			invalidno=len(invalidstud)
			invalidstud =Wear_Invalid_Students.objects.filter(gm_id=gmid)
			context_dict={"stud":invalidstud,"invalidno":invalidno,"gmid":gmid,"item":item}
			return render(request, 'su/invalid_student_table.html',context_dict)
        except Wear.DoesNotExist:
			pass
        try:
            item=Event.objects.get(gm_id=gmid)
            invalidno=len(invalidstud)
            invalidstud =Event_Invalid_Students.objects.filter(gm_id=gmid)
            context_dict={"stud":invalidstud,"invalidno":invalidno,"gmid":gmid,"item":item}
            return render(request, 'su/invalid_student_table.html',context_dict)
        except Event.DoesNotExist:
            print("Oops!")
            pass
    elif (request.user.is_staff) :
		try :
			item =Wear.objects.get(gm_id=gmid)
			coord = Coord.objects.get(cg_id=item.cg_id.cg_id)
			if request.user == coord.user:
				invalidstud = Wear_Invalid_Students.objects.filter(gm_id=gmid)
				invalidno = len(invalidstud)
				context_dict={"stud":invalidstud,"invalidno":invalidno,"gmid":gmid,"item":item}
				return render(request, 'su/invalid_student_table.html',context_dict)
		except Wear.DoesNotExist:
			pass
		try :
			item =Event.objects.get(gm_id=gmid)
			coord = Coord.objects.get(cg_id=item.cg_id.cg_id)
			if request.user == coord.user:
				invalidstud = Event_Invalid_Students.objects.filter(gm_id=gmid)
				invalidno = len(invalidstud)
				context_dict={"stud":invalidstud,"invalidno":invalidno,"gmid":gmid,"item":item}
				return render(request, 'su/invalid_student_table.html',context_dict)
		except Event.DoesNotExist:
			pass
		return HttpResponseRedirect("/su")
    else:
		return HttpResponseRedirect('/su/su/login/')

# # NOTE: Future plans:
# coords can edit item details after registering an item
# review and feedback system
# Cab services
